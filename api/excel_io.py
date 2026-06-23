"""Excel (xlsx) import parsing and template generation.

Project rule: NEVER CSV — imports/exports are Excel or PDF only. Import files are
read with openpyxl; templates are produced via the existing ReportEngine so they
share the app's RTL/Arabic styling.
"""
import io
from typing import List, Dict

from openpyxl import load_workbook


def parse_xlsx(file_bytes: bytes, columns: List[str]) -> List[Dict[str, str]]:
    """Parse an uploaded xlsx into a list of row dicts keyed by ``columns``.

    The first sheet's first row is treated as a header. Columns are matched by the
    Arabic label; if the header doesn't match, columns are taken positionally.
    Fully-empty rows are skipped.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    # Map each requested column to a source index (by label, else by position).
    index_of = {}
    for i, col in enumerate(columns):
        if col in header:
            index_of[col] = header.index(col)
        elif i < len(header):
            index_of[col] = i
        else:
            index_of[col] = None

    out: List[Dict[str, str]] = []
    for raw in rows[1:]:
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        record = {}
        for col in columns:
            idx = index_of[col]
            val = raw[idx] if (idx is not None and idx < len(raw)) else None
            record[col] = "" if val is None else str(val).strip()
        out.append(record)
    wb.close()
    return out


def build_template(report_engine, module_name: str, headers: List[str],
                   title: str = None, sample_row: List[str] = None) -> str:
    """Generate a headers-only (optionally one sample row) xlsx template and
    return its path. Reuses ReportEngine.export_to_excel for consistent styling."""
    data = [sample_row] if sample_row else []
    return report_engine.export_to_excel(module_name, headers, data, title=title)
