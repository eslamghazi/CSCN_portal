from openpyxl import load_workbook

from application.services.report_engine import ReportEngine

HEADERS = ["التاريخ", "النوع", "المبلغ", "البيان"]
ROWS = [
    ["2026-01-01", "إيراد", "1000.00", "منحة"],
    ["2026-02-01", "مصروف", "400.00", "مصروفات"],
]


def test_export_excel_is_rtl_with_header(tmp_path):
    eng = ReportEngine(export_dir=str(tmp_path))
    path = eng.export_to_excel("fin", HEADERS, ROWS, title="تقرير")
    assert path.endswith(".xlsx")
    ws = load_workbook(path).active
    assert ws.sheet_view.rightToLeft is True
    assert ws.cell(1, 1).value == HEADERS[0]
    assert ws.max_row == 3  # header + 2 rows


def test_export_pdf_produces_valid_file(tmp_path):
    eng = ReportEngine(export_dir=str(tmp_path))
    path = eng.export_to_pdf("fin", HEADERS, ROWS, title="تقرير")
    assert path.endswith(".pdf")
    with open(path, "rb") as f:
        assert f.read(5) == b"%PDF-"


def test_export_pdf_handles_empty_data(tmp_path):
    eng = ReportEngine(export_dir=str(tmp_path))
    path = eng.export_to_pdf("fin", HEADERS, [], title="فارغ")
    with open(path, "rb") as f:
        assert f.read(5) == b"%PDF-"
