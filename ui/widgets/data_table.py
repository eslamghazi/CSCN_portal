import math
from datetime import datetime
from typing import List

from qtpy.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
    QTableWidgetItem, QHeaderView, QLabel, QLineEdit, QComboBox, QFileDialog
)
from qtpy.QtCore import Qt, Signal, QSize
from loguru import logger

from ui.themes.colors import Colors
from ui.themes.icons import icon
from ui.widgets.status_badge import StatusBadge
from ui.widgets.icon_button import IconButton
from ui.widgets.cell import cell_widget
from ui.widgets.toast import Toast


class Badge:
    """A status-pill cell value."""
    def __init__(self, text: str, kind: str = "info"):
        self.text = text
        self.kind = kind


class Widget:
    """An arbitrary-widget cell value; `factory` is a callable returning a QWidget
    (rebuilt on each render, e.g. a progress bar)."""
    def __init__(self, factory):
        self.factory = factory


class Action:
    """A row action -> icon-only button in the actions column."""
    def __init__(self, icon_name, tooltip, callback, variant="secondary"):
        self.icon_name = icon_name
        self.tooltip = tooltip
        self.callback = callback
        self.variant = variant


class Row:
    """One table row.

    cells: list of str | Badge for the non-action columns.
    actions: list[Action] rendered in the last ("إجراءات") column.
    search: text matched against the search box.
    sort: per-column sort keys (parallel to cells); falls back to cell text.
    tags: {filter_title: value} matched against filter combos.
    """
    def __init__(self, cells, actions=None, search="", sort=None, tags=None):
        self.cells = cells
        self.actions = actions or []
        self.search = search
        self.sort = sort or []
        self.tags = tags or {}


def _norm(value):
    """Type-stable sort key so str/number/None never compare-clash."""
    if value is None:
        return (2, "")
    if isinstance(value, (int, float)):
        return (0, value)
    return (1, str(value).lower())


class DataTable(QWidget):
    """Reusable table with client-side search, filtering, column sorting,
    pagination and dual empty states. Feed it Row models via set_records()."""

    page_changed = Signal(int)        # kept for backward compatibility
    search_requested = Signal(str)    # kept for backward compatibility
    row_action_requested = Signal(int, str)
    refresh_requested = Signal()      # the toolbar "تحديث" button

    def __init__(self, columns: List[str], page_size: int = 25):
        super().__init__()
        self.columns = columns
        self._all_rows: List[Row] = []
        self._query = ""
        self._filters = []  # list of {"title": str, "combo": QComboBox}
        self._sort_col = None
        self._sort_desc = False
        self._page = 1
        self._page_size = page_size
        self.export_title = "بيانات"
        self._import_columns = None
        self._on_import_row = None
        self._import_sample = None
        self.setup_ui()

    # ------------------------------------------------------------------ UI
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(10)

        self.toolbar = QHBoxLayout()
        search_icon = QLabel()
        si = icon("search", color=Colors.TEXT_MUTED)
        if not si.isNull():
            search_icon.setPixmap(si.pixmap(QSize(16, 16)))
        self.toolbar.addWidget(search_icon)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("بحث...")
        self.search_input.setMaximumWidth(300)
        self.search_input.textChanged.connect(self._on_search_changed)
        self.toolbar.addWidget(self.search_input)
        self.toolbar.addStretch()

        # Tools (right side): import + template (shown when enabled), export, refresh.
        self.import_btn = IconButton("استيراد", "upload", variant="secondary", compact=True)
        self.import_btn.clicked.connect(self._on_import)
        self.import_btn.hide()
        self.template_btn = IconButton("قالب Excel", "browse", variant="secondary", compact=True)
        self.template_btn.clicked.connect(self._on_template)
        self.template_btn.hide()
        self.export_btn = IconButton("تصدير", "export", variant="secondary", compact=True)
        self.export_btn.clicked.connect(self._on_export)
        self.refresh_btn = IconButton("تحديث", "refresh", variant="secondary", compact=True)
        self.refresh_btn.clicked.connect(self.refresh_requested.emit)
        for b in (self.import_btn, self.template_btn, self.export_btn, self.refresh_btn):
            self.toolbar.addWidget(b)
        layout.addLayout(self.toolbar)

        self.table = QTableWidget()
        self.table.setColumnCount(len(self.columns))
        self.table.setHorizontalHeaderLabels(self.columns)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setSortIndicatorShown(True)
        self.table.horizontalHeader().sectionClicked.connect(self._on_sort)
        self.table.verticalHeader().hide()
        self.table.verticalHeader().setDefaultSectionSize(46)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setStyleSheet(
            f"QTableWidget::item:hover {{ background-color: {Colors.PRIMARY_SOFT}; }}"
        )
        layout.addWidget(self.table)

        self.empty_label = QLabel("لا توجد بيانات لعرضها")
        self.empty_label.setAlignment(Qt.AlignCenter)
        self.empty_label.setStyleSheet(
            f"font-size: 16px; color: {Colors.TEXT_MUTED}; margin: 24px;"
        )
        self.empty_label.hide()
        layout.addWidget(self.empty_label)

        pagination = QHBoxLayout()
        self.prev_btn = IconButton("السابق", variant="secondary", compact=True)
        self.prev_btn.clicked.connect(self.prev_page)
        self.page_label = QLabel("صفحة 1 من 1")
        self.page_label.setAlignment(Qt.AlignCenter)
        self.next_btn = IconButton("التالي", variant="secondary", compact=True)
        self.next_btn.clicked.connect(self.next_page)
        pagination.addStretch()
        pagination.addWidget(self.next_btn)
        pagination.addWidget(self.page_label)
        pagination.addWidget(self.prev_btn)
        pagination.addStretch()
        layout.addLayout(pagination)

    def add_filter(self, title: str, options: list):
        """Add a filter combo. options: list of (label, value); a leading
        "الكل" (value None) is added automatically. Rows are matched on
        row.tags[title] == selected value."""
        label = QLabel(title)
        combo = QComboBox()
        combo.addItem("الكل", None)
        for opt_label, value in options:
            combo.addItem(opt_label, value)
        combo.currentIndexChanged.connect(self._on_filter_changed)
        # insert before the trailing stretch
        self.toolbar.insertWidget(self.toolbar.count() - 1, label)
        self.toolbar.insertWidget(self.toolbar.count() - 1, combo)
        self._filters.append({"title": title, "combo": combo})

    # --------------------------------------------------------------- data
    def set_records(self, rows: list):
        self._all_rows = rows or []
        self._page = 1
        self._render()

    def set_data(self, data: list):
        """Backward-compatible: list of list[str] -> plain text rows."""
        rows = [Row(cells=list(r), search=" ".join(str(c) for c in r), sort=list(r))
                for r in data]
        self.set_records(rows)

    # ------------------------------------------------------------- filters
    def _visible(self) -> list:
        rows = self._all_rows
        if self._query:
            q = self._query.lower()
            rows = [r for r in rows if q in (r.search or "").lower()]
        for f in self._filters:
            value = f["combo"].currentData()
            if value is not None:
                rows = [r for r in rows if r.tags.get(f["title"]) == value]
        if self._sort_col is not None:
            col = self._sort_col

            def key(r):
                if col < len(r.sort) and r.sort[col] is not None:
                    return _norm(r.sort[col])
                cell = r.cells[col] if col < len(r.cells) else ""
                return _norm(cell.text if isinstance(cell, Badge) else cell)

            rows = sorted(rows, key=key, reverse=self._sort_desc)
        return rows

    def _render(self):
        visible = self._visible()
        total = len(visible)

        if not self._all_rows:
            self._show_empty("لا توجد بيانات لعرضها")
            return
        if total == 0:
            self._show_empty("لا توجد نتائج مطابقة لبحثك")
            return

        self.table.show()
        self.empty_label.hide()
        self._set_pagination_visible(True)

        pages = max(1, math.ceil(total / self._page_size))
        self._page = min(self._page, pages)
        start = (self._page - 1) * self._page_size
        page_rows = visible[start:start + self._page_size]

        action_col = len(self.columns) - 1
        self.table.setRowCount(len(page_rows))
        for r_idx, row in enumerate(page_rows):
            for c_idx, cell in enumerate(row.cells):
                if c_idx >= len(self.columns):
                    break
                if isinstance(cell, Badge):
                    self.table.setItem(r_idx, c_idx, QTableWidgetItem(""))
                    self.table.setCellWidget(
                        r_idx, c_idx, StatusBadge.for_cell(cell.text, cell.kind))
                elif isinstance(cell, Widget):
                    self.table.setItem(r_idx, c_idx, QTableWidgetItem(""))
                    self.table.setCellWidget(r_idx, c_idx, cell.factory())
                else:
                    item = QTableWidgetItem(str(cell))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.table.setItem(r_idx, c_idx, item)
            if row.actions:
                buttons = [
                    self._make_action_button(a) for a in row.actions
                ]
                self.table.setCellWidget(r_idx, action_col, cell_widget(*buttons))

        self.page_label.setText(f"صفحة {self._page} من {pages}")
        self.prev_btn.setEnabled(self._page > 1)
        self.next_btn.setEnabled(self._page < pages)

    @staticmethod
    def _make_action_button(action: Action) -> IconButton:
        btn = IconButton(icon_name=action.icon_name, variant=action.variant,
                         tooltip=action.tooltip, icon_only=True)
        btn.clicked.connect(lambda _checked=False, cb=action.callback: cb())
        return btn

    def _show_empty(self, message: str):
        self.table.hide()
        self.empty_label.setText(message)
        self.empty_label.show()
        self._set_pagination_visible(False)

    def _set_pagination_visible(self, visible: bool):
        self.prev_btn.setVisible(visible)
        self.next_btn.setVisible(visible)
        self.page_label.setVisible(visible)

    # ------------------------------------------------------------- events
    def _on_search_changed(self, text: str):
        self._query = text.strip()
        self._page = 1
        self._render()

    def _on_filter_changed(self, _index: int):
        self._page = 1
        self._render()

    def _on_sort(self, col: int):
        if self._sort_col == col:
            self._sort_desc = not self._sort_desc
        else:
            self._sort_col = col
            self._sort_desc = False
        order = Qt.DescendingOrder if self._sort_desc else Qt.AscendingOrder
        self.table.horizontalHeader().setSortIndicator(col, order)
        self._render()

    def prev_page(self):
        if self._page > 1:
            self._page -= 1
            self._render()

    def next_page(self):
        self._page += 1
        self._render()

    # --------------------------------------------------------- export/import
    def set_export_title(self, title: str):
        """Name used for the exported file and report heading."""
        self.export_title = title or "بيانات"

    def enable_import(self, columns: list, on_import_row, sample_row=None):
        """Turn on the Import + Template buttons. `columns` are the template
        headers; `on_import_row({header: value})` creates one record per data
        row; `sample_row` is an optional example row written into the template."""
        self._import_columns = list(columns)
        self._on_import_row = on_import_row
        self._import_sample = sample_row
        self.import_btn.show()
        self.template_btn.show()

    def _export_headers_and_rows(self):
        headers = list(self.columns)
        if headers and headers[-1] == "إجراءات":  # drop the actions column
            headers = headers[:-1]
        data = []
        for row in self._all_rows:
            cells = []
            for i in range(len(headers)):
                c = row.cells[i] if i < len(row.cells) else ""
                if isinstance(c, Badge):
                    cells.append(c.text)
                elif isinstance(c, Widget):
                    cells.append("")
                else:
                    cells.append("" if c is None else str(c))
            data.append(cells)
        return headers, data

    def _on_export(self):
        from application.services.report_engine import ReportEngine
        from config.settings import EXPORTS_DIR
        headers, data = self._export_headers_and_rows()
        default = f"{self.export_title} - {datetime.now().strftime('%Y%m%d_%H%M')}"
        dest, selected = QFileDialog.getSaveFileName(
            self, "تصدير الجدول", default, "Excel (*.xlsx);;PDF (*.pdf)")
        if not dest:
            return
        is_pdf = dest.lower().endswith(".pdf") or "pdf" in (selected or "").lower()
        try:
            eng = ReportEngine(export_dir=str(EXPORTS_DIR))
            if is_pdf:
                if not dest.lower().endswith(".pdf"):
                    dest += ".pdf"
                eng.export_to_pdf(self.export_title, headers, data,
                                  title=self.export_title, dest_path=dest)
            else:
                if not dest.lower().endswith(".xlsx"):
                    dest += ".xlsx"
                eng.export_to_excel(self.export_title, headers, data,
                                    title=self.export_title, dest_path=dest)
            Toast.success(self, "تم تصدير الجدول بنجاح")
        except Exception as e:
            logger.error(f"Table export failed: {e}")
            Toast.error(self, "تعذّر تصدير الجدول")

    def _on_template(self):
        if not self._import_columns:
            return
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        default = f"قالب - {self.export_title}.xlsx"
        dest, _ = QFileDialog.getSaveFileName(
            self, "حفظ قالب الاستيراد", default, "Excel (*.xlsx)")
        if not dest:
            return
        if not dest.lower().endswith(".xlsx"):
            dest += ".xlsx"
        try:
            wb = Workbook()
            ws = wb.active
            ws.sheet_view.rightToLeft = True
            ws.append(self._import_columns)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2B4C7E")
            if self._import_sample:
                ws.append(list(self._import_sample))
            for i in range(1, len(self._import_columns) + 1):
                ws.column_dimensions[get_column_letter(i)].width = 22
            wb.save(dest)
            Toast.success(self, "تم حفظ ملف القالب (Excel)")
        except Exception as e:
            logger.error(f"Template export failed: {e}")
            Toast.error(self, "تعذّر حفظ القالب")

    def _on_import(self):
        if not self._on_import_row:
            return
        path, _ = QFileDialog.getOpenFileName(
            self, "استيراد من Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import load_workbook
            ws = load_workbook(path, data_only=True).active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as e:
            logger.error(f"Import read failed: {e}")
            Toast.error(self, "تعذّر قراءة الملف")
            return
        if not rows:
            Toast.error(self, "الملف فارغ")
            return
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        ok = fail = 0
        for r in rows[1:]:
            if r is None or all(v is None for v in r):
                continue
            data = {headers[i]: r[i] for i in range(len(headers)) if i < len(r)}
            try:
                self._on_import_row(data)
                ok += 1
            except Exception as e:
                logger.error(f"Import row failed: {e}")
                fail += 1
        self.refresh_requested.emit()
        if ok:
            Toast.success(self, f"تم استيراد {ok} سجل" + (f" (تعذّر {fail})" if fail else ""))
        else:
            Toast.error(self, "لم يتم استيراد أي سجل")
