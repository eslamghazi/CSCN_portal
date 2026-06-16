from pathlib import Path
from datetime import datetime
from typing import List, Any

from loguru import logger

ORG_NAME = "مركز تنمية مهارات التمريض - جامعة كفر الشيخ"

# Brand colors (hex without '#') used by the exporters.
_PRIMARY = "2B4C7E"
_ALT_ROW = "F8FAFC"
_BORDER = "CBD5E1"


class ReportEngine:
    """Generates reports as Excel (.xlsx) and PDF only — never CSV.

    Excel uses openpyxl (RTL sheet, styled header). PDF uses reportlab with
    Arabic shaping (arabic-reshaper + python-bidi) and an embedded TTF that
    supports Arabic glyphs.
    """

    def __init__(self, export_dir: str = "data/exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        self._pdf_font = None  # cached registered font name

    def _path(self, module_name: str, ext: str) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self.export_dir / f"{module_name}_report_{timestamp}.{ext}"

    # ---------------------------------------------------------------- Excel
    def export_to_excel(self, module_name: str, headers: List[str],
                        data: List[List[Any]], title: str = None,
                        dest_path: str = None) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        file_path = Path(dest_path) if dest_path else self._path(module_name, "xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = (title or module_name)[:31]
        ws.sheet_view.rightToLeft = True

        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color=_BORDER)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.append(list(headers))
        header_fill = PatternFill("solid", fgColor=_PRIMARY)
        header_font = Font(bold=True, color="FFFFFF", size=12)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[1].height = 24

        for row in data:
            ws.append(list(row))
        alt_fill = PatternFill("solid", fgColor=_ALT_ROW)
        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.alignment = center
                cell.border = border
                if r_idx % 2 == 0:
                    cell.fill = alt_fill

        for col_idx, header in enumerate(headers, start=1):
            values = [str(header)] + [
                str(row[col_idx - 1]) for row in data if col_idx - 1 < len(row)
            ]
            width = min(max((len(v) for v in values), default=10) + 4, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"
        wb.save(file_path)
        return str(file_path.absolute())

    # ------------------------------------------------------------------ PDF
    @staticmethod
    def _shape(text: Any) -> str:
        """Reshape + bidi-reorder Arabic text for correct PDF rendering."""
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display
            return get_display(arabic_reshaper.reshape(str(text)))
        except Exception:
            return str(text)

    def _register_pdf_font(self) -> str:
        """Register a TTF that supports Arabic and return its name. Looks in
        bundled fonts first, then common Windows fonts; falls back to Helvetica
        (Latin only) if none is found."""
        if self._pdf_font:
            return self._pdf_font

        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        candidates = []
        res = Path("ui/resources/fonts")
        if res.exists():
            candidates += sorted(res.glob("*.ttf"))
        win = Path("C:/Windows/Fonts")
        candidates += [win / n for n in ("tahoma.ttf", "arial.ttf", "segoeui.ttf")]

        for font_path in candidates:
            try:
                if font_path.exists():
                    pdfmetrics.registerFont(TTFont("AppArabic", str(font_path)))
                    self._pdf_font = "AppArabic"
                    return self._pdf_font
            except Exception as e:
                logger.warning(f"Could not register PDF font {font_path}: {e}")
        logger.warning("No Arabic-capable TTF found; PDF Arabic text may not render.")
        self._pdf_font = "Helvetica"
        return self._pdf_font

    def export_to_pdf(self, module_name: str, headers: List[str],
                      data: List[List[Any]], title: str = None,
                      dest_path: str = None) -> str:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        )

        file_path = Path(dest_path) if dest_path else self._path(module_name, "pdf")
        font = self._register_pdf_font()

        doc = SimpleDocTemplate(
            str(file_path), pagesize=landscape(A4),
            rightMargin=15 * mm, leftMargin=15 * mm,
            topMargin=16 * mm, bottomMargin=15 * mm,
            title=str(title or module_name),
        )
        base = getSampleStyleSheet()
        org_style = ParagraphStyle(
            "Org", parent=base["Normal"], fontName=font, alignment=TA_CENTER,
            fontSize=10, textColor=colors.HexColor("#566375"))
        title_style = ParagraphStyle(
            "TitleAr", parent=base["Title"], fontName=font, alignment=TA_CENTER,
            fontSize=18, spaceBefore=4, spaceAfter=2,
            textColor=colors.HexColor("#2B4C7E"))
        meta_style = ParagraphStyle(
            "Meta", parent=base["Normal"], fontName=font, alignment=TA_CENTER,
            fontSize=9, textColor=colors.HexColor("#8A97A8"))

        story = [
            Paragraph(self._shape(ORG_NAME), org_style),
            Paragraph(self._shape(title or module_name), title_style),
            Paragraph(self._shape(
                f"تاريخ التوليد: {datetime.now().strftime('%Y-%m-%d %H:%M')}"), meta_style),
            Spacer(1, 10),
        ]

        # Reverse columns so the first logical column appears on the right (RTL).
        table_rows = [[self._shape(h) for h in headers][::-1]]
        for row in data:
            table_rows.append([self._shape(c) for c in row][::-1])

        table = Table(table_rows, repeatRows=1)
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B4C7E")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(table)
        if not data:
            story.append(Spacer(1, 8))
            story.append(Paragraph(self._shape("لا توجد بيانات لعرضها."), meta_style))

        doc.build(story)
        return str(file_path.absolute())
